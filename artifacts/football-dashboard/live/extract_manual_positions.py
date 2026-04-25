"""
extract_manual_positions.py
---------------------------
Extrait les positions des joueurs Top 5 du fichier Excel "Buteurs Maison 4.1"
(onglet "Joueurs", colonne "Poste") et produit `live/data/manual_positions.json`.

Le fichier Excel contient ~13 650 joueurs avec position fine annotée à la main
(BU/AT/AD/AG/MOC/MDC/MC/DC/DD/DG/G…), ce qui constitue une source de vérité
nettement plus fiable que les champs `position`/`specific_position` BSD
(grossiers à 82 % pour les joueurs Top 5).

Usage
-----
    cd artifacts/football-dashboard
    python -m live.extract_manual_positions \\
        --xlsx attached_assets/Buteurs_Maison_4.1_1777117790057.xlsx
    # → live/data/manual_positions.json + rapport sur stdout

Format du JSON produit
----------------------
{
  "by_key":      { "<name_norm>|<team_norm>|<country>": "ST", ... },
  "by_name":     { "<name_norm>": [{"team": "...", "country": "ESP", "poste": "SS"}, ...] },
  "metadata":    { "source": "...", "extracted_at": "...", "n_total": ..., "n_top5": ... }
}

Mapping postes Excel (FR) → codes BSD (EN)
------------------------------------------
    BU  → ST       (Buteur)
    AT  → SS       (Attaquant retiré / 2nd attaquant — type Griezmann)
    AC  → CF       (Attaquant central — rare)
    AD  → RW       (Ailier droit)
    AG  → LW       (Ailier gauche)
    MOC → AM       (Milieu offensif central)
    MD  → RM       (Milieu droit)
    MG  → LM       (Milieu gauche)
    MC  → CM       (Milieu central)
    MDC → DM       (Milieu défensif central)
    DD  → RB       (Défenseur droit)
    DG  → LB       (Défenseur gauche)
    DC  → CB       (Défenseur central)
    G   → GK       (Gardien)
"""
from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")


# === Mapping poste FR → code BSD ============================================
POSTE_MAP = {
    "BU": "ST", "AT": "SS", "AC": "CF",
    "AD": "RW", "AG": "LW",
    "MOC": "AM", "MD": "RM", "MG": "LM",
    "MC": "CM", "MDC": "DM",
    "DD": "RB", "DG": "LB", "DC": "CB",
    "G": "GK",
}


# === Mapping nom équipe Excel (FR) → nom équipe BSD (EN/local) ==============
# Construit manuellement à partir de l'analyse des squads cache BSD.
# Couvre les 96 équipes Top 5 (saison 2025-2026). À mettre à jour si BSD
# change ses libellés ou si une nouvelle équipe arrive (promotion).
TEAM_NAME_MAP: dict[str, dict[str, str]] = {
    "ENG": {
        "Brighton": "Brighton & Hove Albion",
        "Leeds": "Leeds United",
        "Newcastle": "Newcastle United",
        "Tottenham": "Tottenham Hotspur",
        "West Ham": "West Ham United",
    },
    "ESP": {
        "Alaves": "Deportivo Alavés",
        "Athletic Bilbao": "Athletic Club",
        "Atletico Madrid": "Atlético Madrid",
        "Barcelone": "FC Barcelona",
        "Betis": "Real Betis",
        "Girona": "Girona FC",
        "Levante": "Levante UD",
        "Majorque": "Mallorca",
        "Seville": "Sevilla",
        "Valence": "Valencia",
    },
    "ITA": {
        "AS Rome": "Roma",
        "Bologne": "Bologna",
        "Hellas Verone": "Hellas Verona",
        "Inter Milan": "Inter",
        "Lazio Rome": "Lazio",
        "Milan AC": "Milan",
        "Naples": "Napoli",
        "Parme": "Parma",
    },
    "GER": {
        "Augsbourg": "FC Augsburg",
        "Bayer Leverkusen": "Bayer 04 Leverkusen",
        "Bayern Munich": "FC Bayern München",
        "Cologne": "1. FC Köln",
        "Dortmund": "Borussia Dortmund",
        "Francfort": "Eintracht Frankfurt",
        "Fribourg": "SC Freiburg",
        "Hambourg": "Hamburger SV",
        "Heidenheim": "1. FC Heidenheim",
        "Hoffenheim": "TSG Hoffenheim",
        "Leipzig": "RB Leipzig",
        "M'Gladbach": "Borussia M'gladbach",
        "Mayence": "1. FSV Mainz 05",
        "St. Pauli": "FC St. Pauli",
        "Stuttgart": "VfB Stuttgart",
        "Union Berlin": "1. FC Union Berlin",
        "Werder Brême": "SV Werder Bremen",
        "Wolfsbourg": "VfL Wolfsburg",
    },
    "FRA": {
        "Brest": "Stade Brestois",
        "Lens": "RC Lens",
        "Lyon": "Olympique Lyonnais",
        "Marseille": "Olympique de Marseille",
        "Monaco": "AS Monaco",
        "Paris SG": "Paris Saint-Germain",
        "Rennes": "Stade Rennais",
        "Strasbourg": "RC Strasbourg",
    },
}

TOP5_COUNTRIES = {"ENG", "ESP", "ITA", "GER", "FRA"}

DEFAULT_XLSX = Path("../../attached_assets/Buteurs_Maison_4.1_1777117790057.xlsx")
DEFAULT_OUT = Path("live/data/manual_positions.json")


def normalize(s: str | None) -> str:
    """Lowercase + strip accents + strip ponctuation utile au matching."""
    if not s:
        return ""
    s = str(s).strip()
    nfkd = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    s = s.lower()
    # Suppress ponctuation usuelle
    for ch in (".", ",", "'", "-"):
        s = s.replace(ch, " ")
    return " ".join(s.split())


def map_team_excel_to_bsd(country: str, team_excel: str) -> str:
    """Retourne le nom BSD attendu (ou le nom Excel si déjà identique)."""
    return TEAM_NAME_MAP.get(country, {}).get(team_excel, team_excel)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                    help="Chemin vers le fichier Buteurs_Maison_x.x.xlsx")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help="Chemin du JSON de sortie")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"❌ Fichier introuvable : {args.xlsx}")

    print(f"📂 Lecture {args.xlsx} …", flush=True)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=False)
    if "Joueurs" not in wb.sheetnames:
        sys.exit(f"❌ Onglet 'Joueurs' absent. Sheets: {wb.sheetnames}")
    ws = wb["Joueurs"]

    # Header parsing (col index based on observed file)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    needed = {"Joueur", "Equipe", "Pays", "Poste"}
    missing = needed - set(col)
    if missing:
        sys.exit(f"❌ Colonnes manquantes : {missing}. Headers: {headers}")

    by_key: dict[str, str] = {}
    by_name: dict[str, list[dict]] = defaultdict(list)
    n_total = 0
    n_top5 = 0
    n_skipped_no_poste = 0
    n_unknown_poste: dict[str, int] = defaultdict(int)
    countries_count: dict[str, int] = defaultdict(int)

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, col["Joueur"]).value
        eq = ws.cell(r, col["Equipe"]).value
        py = ws.cell(r, col["Pays"]).value
        poste_fr = ws.cell(r, col["Poste"]).value
        if not nm:
            continue
        n_total += 1
        if py not in TOP5_COUNTRIES:
            continue
        n_top5 += 1
        countries_count[py] += 1
        if not poste_fr:
            n_skipped_no_poste += 1
            continue
        poste_fr = str(poste_fr).strip().upper()
        if poste_fr not in POSTE_MAP:
            n_unknown_poste[poste_fr] += 1
            continue
        poste_bsd = POSTE_MAP[poste_fr]

        team_bsd = map_team_excel_to_bsd(py, str(eq).strip()) if eq else ""
        nm_norm = normalize(nm)
        team_norm = normalize(team_bsd)
        key = f"{nm_norm}|{team_norm}|{py}"
        by_key[key] = poste_bsd
        by_name[nm_norm].append({"team": team_bsd, "country": py, "poste": poste_bsd})

    payload = {
        "by_key": by_key,
        "by_name": dict(by_name),
        "metadata": {
            "source": str(args.xlsx.name),
            "extracted_at": datetime.now(timezone.utc).isoformat(),
            "n_total_excel": n_total,
            "n_top5": n_top5,
            "n_extracted": len(by_key),
            "n_skipped_no_poste": n_skipped_no_poste,
            "n_unknown_poste": dict(n_unknown_poste),
            "countries_count": dict(countries_count),
            "poste_map": POSTE_MAP,
            "team_name_aliases": {c: list(v.keys()) for c, v in TEAM_NAME_MAP.items()},
        },
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2))

    print()
    print(f"✅ {len(by_key)} positions extraites Top 5 ({len(by_name)} noms uniques)")
    print(f"   Total Excel : {n_total} joueurs (toutes ligues confondues)")
    print(f"   Top 5      : {n_top5}")
    print(f"   Sans poste : {n_skipped_no_poste}")
    if n_unknown_poste:
        print(f"   Postes inconnus : {dict(n_unknown_poste)}")
    print(f"   Par pays   : {dict(countries_count)}")
    print(f"📝 Écrit dans {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
